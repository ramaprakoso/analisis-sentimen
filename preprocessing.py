import re,string
import nltk
import xlrd
import openpyxl
import csv
from bs4 import BeautifulSoup
from nltk.tokenize import word_tokenize
"""
def fileprocess(filenameInput='tweet.xlsx',filenamePreprocess='data_pre.xlsx'):
    # buka file input
    fileTrain = xlrd.open_workbook(filenameInput)
    dataTrain = fileTrain.sheet_by_index(0)
    rowLen = dataTrain.nrows

    # siapkan file output
    filePreprocessed = openpyxl.Workbook()
    dataPreprocessed = filePreprocessed.active

    # untuk setiap data input, lakukan preprocessing
    # dan hasil preprocessing simpan ke dalam data output
    for i in range(rowLen):
        data_i = dataTrain.cell(i,0).value
        class_i = dataTrain.cell(i, 1).value
        prep = preprocess(data_i)

        if prep:
            for i in range(len(prep)):
                dataPreprocessed.append([''.join(prep[i]), class_i])
    # simpan file output
    filePreprocessed.save(filenamePreprocess)
    return dataPreprocessed
"""
def preprocess(tweet):
	#hapus_html 
	#tweet=BeautifulSoup(tweet)
	#hapus url 
	tweet=re.sub(r'http\S+','',tweet)
	#hapus @username
	tweet=re.sub('@[^\s]+','',tweet)
	#hapus #tagger 
	tweet = re.sub(r'#([^\s]+)', r'\1', tweet)
	#hapus tanda baca
	tweet=hapus_tanda(tweet)
	#hapus angka dan angka yang berada dalam string 
	tweet=re.sub(r'\w*\d\w*', '',tweet).strip()
	#hapus repetisi karakter 
	tweet=hapus_katadouble(tweet)
	
	return tweet   
def hapus_tanda(tweet): 
	tanda_baca = set(string.punctuation)
	tweet = ''.join(ch for ch in tweet if ch not in tanda_baca)
	return tweet
def hapus_katadouble(s): 
    #look for 2 or more repetitions of character and replace with the character itself
    pattern = re.compile(r"(.)\1{1,}", re.DOTALL)
    return pattern.sub(r"\1\1", s)
def hapus_html(tweet):
	pass
def tokenize(tweet): 
	token=word_tokenize(tweet)
	return token 
def tokenize_freq(tweet): 
	words=nltk.tokenize.word_tokenize(tweet)
	fdist=FreqDist(words)
	return fdist
def get_fitur(tweet): 
	#token 
	tokens=tokenize(tweet)
	#tokens=kbbi(tokens)
	tokens=stopwordDel(tokens)
	dasar=[line.strip('\n')for line in open('kamus/rootword.txt')]
	nonKataDasar=list(set(tokens)-set(dasar))
	kataDasar=list(set(tokens)-set(nonKataDasar))
	for i in range(0,len(nonKataDasar)): 
		nonKataDasar[i]=stemming(nonKataDasar[i],dasar)
	tokens=list(set(nonKataDasar+kataDasar))
	tokens=kbbi(tokens)
	#tokens=filter_fitur(tokens)
	return tokens 
def fitur_ekstraksi(inpTweet): 
	tweets=[]
	for row in inpTweet: 
		tweet=row[0]
		label=row[1]
		pre_tweet=preprocess(tweet)
		featureVector=get_fitur(pre_tweet)
		tweets.append((featureVector,label))
	return tweets
def fitur_mentah(tweet): 
	tweets=[]
	for word in tweet: 
		pre=preprocess(word)
		fitur=get_fitur(pre)
		tweets.append(fitur)
	return tweets
def list_negatif(): 
	negatif=[word.strip('\n').strip('\r') for word in open('kamus/negatif_ta.txt')]
	return negatif 
def list_positif():
	positif=[word.strip('\n').strip('\r') for word in open('kamus/positif_ta.txt')]
	return positif
def filter_negatif():
	pass
def filter_positif(): 
	pass
def stopwordDel(token):
	stopword=[word.strip('\n') for word in open('kamus/stopword.txt')] 
	noise=[noise.strip('\n').strip('\r') for noise in open('kamus/noise.txt')]
	tampung=[]
	for i in range(0,len(token)): 
		if token[i] not in stopword and token[i] not in noise: 
			tampung.append(token[i])
	return tampung
def filter_fitur(token): 
	negatif=[word.strip('\n').strip('\r') for word in open('kamus/negatif_ta.txt')]
	positif=[word.strip('\n').strip('\r') for word in open('kamus/positif_ta.txt')]
	t=[]
	for i in range(0,len(token)): 
		if token[i] in negatif or token[i] in positif: 
			t.append(token[i])
	return t
def kbbi(token): 
	kbba=[kamus.strip('\n').strip('\r') for kamus in open('kamus/kbba.txt')]
	#ubah list menjadi dictionary 
	dic={}
	for i in kbba: 
		(key,val)=i.split('\t')
		dic[str(key)]=val
	#kbbi cocokan 
	final_string = ' '.join(str(dic.get(word, word)) for word in token).split()
	return final_string
#stemming kata 
def akhiranPertama(token): 
	#menghapus -kah,-lah,-tah,-pun
	if re.search('(a-z0-9)+kah$',token): 
		token=re.sub(r'(a-zO-9)+kah$',r'\1',token)
		return token 
	if re.search('(a-z0-9)+lah$',token): 
		token=re.sub(r'(a-zO-9)+lah$',r'\1',token)
		return token 
	if re.search('(a-z0-9)+tah$',token): 
		token=re.sub(r'(a-zO-9)+tah$',r'\1',token)
		return token 
	if re.search('(a-z0-9)+pun$',token): 
		token=re.sub(r'(a-zO-9)+pun$',r'\1',token)
		return token 
	return token 
def akhiranKedua(token): 
	#fungsi stemming menghapus kata ganti milik -ku,-mu,-ya
	if re.search('([a-z0-9]+)nya$',token):
		token=re.sub(r'([a-z0-9]+)nya$','r\1',token)
		return token 
	if re.search('([a-z0-9]+)ku$',token):
		token=re.sub(r'([a-z0-9]+)ku$','r\1',token)
		return token 
	if re.search('([a-z0-9]+)mu$',token):
		token=re.sub(r'([a-z0-9]+)mu$','r\1',token)
		return token 
	return token 
def akhiranKetiga(token): 
	if re.search(r'([a-z0-9]+)kan$',token): 
		token=re.sub(r'([a-z0-9]+)kan$','r\1',token)
		return token 
	if re.search(r'([a-z0-9]+)an$',token): 
		token=re.sub(r'([a-z0-9]+)an$','r\1',token)
		return token 
	if re.search(r'([a-z0-9]+)i$',token): 
		token=re.sub(r'([a-z0-9]+)i$','r\1',token)
		return token 
	return token 
def awalanPertama(token):
	#menghapus awalan me-,pe-,di-,ter-,ke-
	if re.search('^meng',token): 
		token=re.sub('^meng','',token)
		return token 
	if re.search('^meny',token):
		 token=re.sub('^meny','s',token)
		 return token 
	if re.search('^men',token): 
		token=re.sub('^men','',token) 
		return token
	if re.search('^me',token): 
		token=re.sub('^me','',token)
		if re.search('^m(^[aiueo])',token): 
			token=re.sub('^m([^aiueo])',r'\1',token)
		return token
	if re.search('^peng',token): 
		token=re.sub('^peng','',token)
		return token 
	if re.search('^peny',token): 
		token=re.sub('^peny','s',token)
		return token 
	if re.search('^pen',token): 
		token=re.sub('^pen','',token)
		return token 
	if re.search('^pem',token): 
		if re.search('^pem[aiueo]',token): 
			token=re.sub('^pem','p',token)
		else: 
			token=re.sub('^pem','',token)
		return token 
	if re.search('^di',token):
		token=re.sub('^di','',token)
		return token 
	if re.search('^ter',token): 
		token=re.sub('^ter','',token)
		return token 
	if re.search('^ke',token): 
		token=re.sub('^ke','',token)
		return token 
	return token 
def awalanKedua(token): 
	if re.search('^ber',token): 
		token=re.sub('^ber','',token)
		return token 
	if re.search('^belajar',token):
		token=re.sub('^belajar','ajar',token)
		return token
	if re.search('^berhasil',token):
		token=re.sub('^berhasil','berhasil',token)
		return token	
	if re.search('^pensiun',token):
		token=re.sub('^pensiun','pensiunmen',token)
		return token  
	#if re.search('^bek+er',token): 
	#	token=re.sub('^bek+er','ker',token)
	#	return token 
	if re.search('^per',token):
		token=re.sub('^per','',token)
		return token 
	if re.search('^pelajar',token): 
		token=re.sub('^pelajar','ajar',token)
		return token 
	if re.search('^pe',token): 
		token=re.sub('^pe','',token)
		return token
	return token  
def stemming(token,dasar):
		#menghapus akhiran -kah, -lah, -tahs
		token=akhiranPertama(token)
		if token in dasar: 
			return token 
		#mengganti kepemilikan kata -ku -mu -nya 
		token=akhiranKedua(token)
		if token in dasar: 
			return token 
			
		tempToken=token 
		token=awalanPertama(token)
		if token in dasar: 
			return token 
		if tempToken==token: 
			token=awalanKedua(token)
			if token in dasar: 
				return token 
			token=akhiranKetiga(token)
			if token in dasar: 
				return token 
		else: 
			temptoken2 = token
			token = akhiranKetiga(token)
			if token in dasar :
				return token
			if token != temptoken2:
				token = awalanKedua(token)
				if token in dasar :
					return token
		return token 		
		
"""			
if __name__ == '__main__':
	fp = open('tweet3000.csv', 'r')
	line = fp.read()
	inpTweets = csv.reader(open('twiit.csv', 'rb'), delimiter=';', quotechar='|')
	pre=preprocess(line)
	fitur=get_fitur(pre)
	negatif=filter_negatif()
	positif=filter_positif()
	print positif
	#print len(positif)
	#print len(negatif)
	#print fitur
	#print line
	#print preprocess(line)
	#print fitur_mentah(line)
	#print fitur_ekstraksi(inpTweets)
	#print fiturlist_negatif()
	#print len(sorted(negatif))
	#print get_svm(inpTweets,negatif)
	
"""
